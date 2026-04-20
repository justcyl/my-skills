#!/usr/bin/env python3
"""
FAIR-style Experiment Tracker Generator
基于凯明（Kaiming He）在 FAIR 的实验管理方法论
生成结构化的实验追踪 Excel 工作簿
"""

import argparse
import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, DataBarRule

# ── 颜色系统 ──
COLORS = {
    "header_bg": "1B2A4A",       # 深蓝 header 背景
    "header_font": "FFFFFF",     # 白色 header 文字
    "baseline_bg": "E8F5E9",     # 浅绿色 baseline 行
    "group_a_bg": "EDE7F6",      # 浅紫色 实验组 A
    "group_b_bg": "E3F2FD",      # 浅蓝色 实验组 B
    "group_c_bg": "FFF3E0",      # 浅橙色 实验组 C
    "failed_bg": "FFEBEE",       # 浅红色 失败/废弃
    "prediction_bg": "FFFDE7",   # 浅黄色 预测列
    "surprise_yes": "FF6F00",    # 橙色 surprise 标记
    "surprise_no": "4CAF50",     # 绿色 no surprise
    "insight_bg": "F3E5F5",      # 浅紫色 insight 列
    "delta_positive": "2E7D32",  # 绿色 正向提升
    "delta_negative": "C62828",  # 红色 负向下降
    "meta_bg": "F5F5F5",         # 浅灰色 meta 信息
    "section_border": "90A4AE",  # 区域分隔线
    "dashboard_title": "0D47A1", # 深蓝色 dashboard 标题
    "dashboard_card_bg": "F8F9FA",
    "dashboard_accent": "1565C0",
}

# ── 领域预设 ──
DOMAIN_METRICS = {
    "cv": {
        "primary": ["Top-1 Acc (%)", "Top-5 Acc (%)", "mAP (%)"],
        "secondary": ["FLOPs (G)", "Params (M)", "Throughput (img/s)"],
        "aux": ["Train Loss", "Val Loss", "Best Epoch"],
    },
    "nlp": {
        "primary": ["PPL ↓", "BLEU ↑", "ROUGE-L ↑"],
        "secondary": ["F1 (%)", "Latency (ms)", "Params (M)"],
        "aux": ["Train Loss", "Val Loss", "Best Epoch"],
    },
    "rl": {
        "primary": ["Avg Return ↑", "Success Rate (%)", "Sample Eff."],
        "secondary": ["Env Steps (M)", "Wall Time (h)", "Params (M)"],
        "aux": ["Train Reward", "Eval Reward", "Best Epoch"],
    },
    "multimodal": {
        "primary": ["Acc (%)", "CIDEr ↑", "CLIPScore ↑"],
        "secondary": ["FLOPs (G)", "Params (M)", "Latency (ms)"],
        "aux": ["Train Loss", "Val Loss", "Best Epoch"],
    },
    "general": {
        "primary": ["Metric-1 ↑", "Metric-2 ↑", "Metric-3 ↑"],
        "secondary": ["FLOPs (G)", "Params (M)", "Speed"],
        "aux": ["Train Loss", "Val Loss", "Best Epoch"],
    },
}

EXPERIMENT_VARIABLES = {
    "ablation": ["Component Removed/Changed", "Variant Detail"],
    "hyperparam": ["Learning Rate", "Batch Size", "Weight Decay", "Scheduler"],
    "architecture": ["Model Name", "Depth", "Width", "Head Count"],
    "method": ["Method Name", "Key Difference"],
    "dataset": ["Dataset", "Train Size", "Augmentation"],
}

# ── 通用样式 ──
THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

SECTION_RIGHT_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="medium", color=COLORS["section_border"]),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

HEADER_FONT = Font(name="Arial", bold=True, color=COLORS["header_font"], size=11)
HEADER_FILL = PatternFill("solid", fgColor=COLORS["header_bg"])
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

BODY_FONT = Font(name="Arial", size=10)
BODY_ALIGN = Alignment(vertical="center", wrap_text=True)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")


def apply_header(ws, row, col, value, width=16):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = HEADER_ALIGN
    cell.border = THIN_BORDER
    letter = get_column_letter(col)
    ws.column_dimensions[letter].width = width
    return cell


def apply_body(ws, row, col, value="", fill_color=None, font_color=None, bold=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name="Arial", size=10, color=font_color or "000000", bold=bold)
    cell.alignment = BODY_ALIGN
    cell.border = THIN_BORDER
    if fill_color:
        cell.fill = PatternFill("solid", fgColor=fill_color)
    return cell


def apply_section_border(ws, row, col):
    """给区域最右列加重边框作为分隔"""
    cell = ws.cell(row=row, column=col)
    cell.border = SECTION_RIGHT_BORDER


def create_dashboard(wb, project_name, domain, exp_type):
    """创建 Dashboard sheet"""
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_properties.tabColor = COLORS["dashboard_accent"]

    # 标题
    ws.merge_cells("B2:H2")
    title_cell = ws.cell(row=2, column=2, value=f"🔬 {project_name} — Experiment Dashboard")
    title_cell.font = Font(name="Arial", size=18, bold=True, color=COLORS["dashboard_title"])
    title_cell.alignment = Alignment(horizontal="left", vertical="center")

    # 项目信息卡片
    info_items = [
        ("Domain", domain.upper()),
        ("Experiment Type", exp_type.replace("_", " ").title()),
        ("Created", datetime.date.today().strftime("%Y-%m-%d")),
        ("Status", "In Progress"),
    ]
    for i, (label, val) in enumerate(info_items):
        r = 4 + i
        lbl_cell = ws.cell(row=r, column=2, value=label)
        lbl_cell.font = Font(name="Arial", size=10, bold=True, color="666666")
        val_cell = ws.cell(row=r, column=3, value=val)
        val_cell.font = Font(name="Arial", size=10, color="333333")

    # 快速参考区
    ws.merge_cells("B9:H9")
    ref_title = ws.cell(row=9, column=2, value="📋 实验管理核心原则（来自 FAIR 方法论）")
    ref_title.font = Font(name="Arial", size=13, bold=True, color=COLORS["dashboard_title"])

    principles = [
        "1. 跑实验前先预测——预测对了继续推进，预测错了就是学习信号",
        "2. 每一行都要和其他行形成对照——对照产生梯度信号",
        "3. 精选指标和实验——不是所有结果都值得记录",
        "4. 避免两个极端：实验太少（无信号）vs 盲目堆实验（无思考）",
        "5. 表格要能讲故事——从上到下应该是一条清晰的探索路径",
    ]
    for i, p in enumerate(principles):
        cell = ws.cell(row=11 + i, column=2, value=p)
        cell.font = Font(name="Arial", size=10, color="444444")
        ws.merge_cells(f"B{11+i}:H{11+i}")

    # 状态汇总（链接到主表的公式）
    ws.merge_cells("B18:H18")
    summary_title = ws.cell(row=18, column=2, value="📊 实验进度概览")
    summary_title.font = Font(name="Arial", size=13, bold=True, color=COLORS["dashboard_title"])

    summary_labels = ["Total Experiments", "Completed", "Running", "Surprises Found"]
    for i, label in enumerate(summary_labels):
        r = 20 + i
        lbl = ws.cell(row=r, column=2, value=label)
        lbl.font = Font(name="Arial", size=10, bold=True)
        val = ws.cell(row=r, column=3, value=0)
        val.font = Font(name="Arial", size=14, bold=True, color=COLORS["dashboard_accent"])

    # 关键发现区域
    ws.merge_cells("B26:H26")
    findings_title = ws.cell(row=26, column=2, value="💡 关键发现 / Key Findings")
    findings_title.font = Font(name="Arial", size=13, bold=True, color=COLORS["dashboard_title"])

    for i in range(5):
        r = 28 + i
        ws.cell(row=r, column=2, value=f"Finding {i+1}:")
        ws.cell(row=r, column=2).font = Font(name="Arial", size=10, bold=True, color="888888")
        ws.merge_cells(f"C{r}:H{r}")
        filler = ws.cell(row=r, column=3, value="[在此记录关键发现]")
        filler.font = Font(name="Arial", size=10, color="BBBBBB", italic=True)

    # 列宽
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 30
    for c in "DEFGH":
        ws.column_dimensions[c].width = 14

    ws.sheet_view.showGridLines = False


def create_main_experiments(wb, domain, exp_type, baseline_name, num_rows):
    """创建主实验表格 sheet"""
    ws = wb.create_sheet("Main Experiments")
    ws.sheet_properties.tabColor = "1565C0"

    metrics = DOMAIN_METRICS.get(domain, DOMAIN_METRICS["general"])
    variables = EXPERIMENT_VARIABLES.get(exp_type, EXPERIMENT_VARIABLES["ablation"])

    # ── 构建列定义 ──
    # 区域: Meta | Variables | Prediction | Results (Primary) | Results (Secondary) | Results (Aux) | Analysis
    sections = []

    # Meta 列
    meta_cols = [
        ("Exp ID", 10), ("Date", 12), ("Description", 30),
        ("Hypothesis", 30), ("Status", 12),
    ]
    sections.append(("META", meta_cols))

    # 变量列
    var_cols = [(v, 18) for v in variables]
    sections.append(("VARIABLES", var_cols))

    # 预测列
    pred_cols = [("Prediction", 24), ("Confidence", 12)]
    sections.append(("PREDICTION", pred_cols))

    # 主指标
    pri_cols = [(m, 14) for m in metrics["primary"]]
    sections.append(("PRIMARY METRICS", pri_cols))

    # 辅助指标
    sec_cols = [(m, 14) for m in metrics["secondary"]]
    sections.append(("SECONDARY METRICS", sec_cols))

    # Aux 指标
    aux_cols = [(m, 14) for m in metrics["aux"]]
    sections.append(("AUX METRICS", aux_cols))

    # 分析列
    analysis_cols = [
        ("Δ vs Baseline", 14), ("Surprise?", 12),
        ("Insight", 35), ("Next Action", 30),
    ]
    sections.append(("ANALYSIS", analysis_cols))

    # Config 列
    config_cols = [("Config / Command", 40), ("Notes", 30)]
    sections.append(("CONFIG", config_cols))

    # ── 第一行：Section 标题 ──
    col = 1
    section_ranges = {}
    for section_name, cols in sections:
        start_col = col
        end_col = col + len(cols) - 1
        section_ranges[section_name] = (start_col, end_col)
        if len(cols) > 1:
            ws.merge_cells(
                start_row=1, start_column=start_col,
                end_row=1, end_column=end_col
            )
        section_cell = ws.cell(row=1, column=start_col, value=section_name)
        section_cell.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
        section_cell.fill = PatternFill("solid", fgColor="37474F")
        section_cell.alignment = Alignment(horizontal="center", vertical="center")
        # 填充合并区域
        for c in range(start_col, end_col + 1):
            ws.cell(row=1, column=c).fill = PatternFill("solid", fgColor="37474F")
            ws.cell(row=1, column=c).font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
        col = end_col + 1

    # ── 第二行：列标题 ──
    col = 1
    for section_name, cols in sections:
        for col_name, width in cols:
            apply_header(ws, 2, col, col_name, width)
            col += 1

    # ── 区域右边框 ──
    for section_name, (start_c, end_c) in section_ranges.items():
        for r in range(1, 3 + num_rows + 5):
            apply_section_border(ws, r, end_c)

    # ── 数据行 ──
    total_cols = sum(len(cols) for _, cols in sections)

    # 找到各区域列号
    pred_start = section_ranges["PREDICTION"][0]
    pred_end = section_ranges["PREDICTION"][1]
    primary_start = section_ranges["PRIMARY METRICS"][0]
    analysis_start = section_ranges["ANALYSIS"][0]
    surprise_col = analysis_start + 1  # Surprise? 列
    insight_col = analysis_start + 2   # Insight 列
    delta_col = analysis_start         # Δ vs Baseline 列

    # Baseline 行（第 3 行）
    baseline_row = 3
    for c in range(1, total_cols + 1):
        bg = COLORS["baseline_bg"]
        if pred_start <= c <= pred_end:
            bg = COLORS["prediction_bg"]
        apply_body(ws, baseline_row, c, fill_color=bg)

    ws.cell(row=baseline_row, column=1, value="E000")
    ws.cell(row=baseline_row, column=1).font = Font(name="Arial", size=10, bold=True, color="2E7D32")
    ws.cell(row=baseline_row, column=3, value=f"Baseline: {baseline_name}")
    ws.cell(row=baseline_row, column=3).font = Font(name="Arial", size=10, bold=True)
    ws.cell(row=baseline_row, column=section_ranges["META"][0] + 4,
            value="Done")

    # Δ vs Baseline 列 baseline 行写 "—"
    ws.cell(row=baseline_row, column=delta_col, value="—")
    ws.cell(row=baseline_row, column=delta_col).alignment = CENTER_ALIGN

    # 实验行
    group_colors = [COLORS["group_a_bg"], COLORS["group_b_bg"], COLORS["group_c_bg"]]
    for i in range(num_rows):
        r = baseline_row + 1 + i
        exp_id = f"E{i+1:03d}"
        group_idx = i // 4  # 每 4 个一组
        bg = group_colors[group_idx % len(group_colors)]

        for c in range(1, total_cols + 1):
            cell_bg = bg
            if pred_start <= c <= pred_end:
                cell_bg = COLORS["prediction_bg"]
            elif c == insight_col:
                cell_bg = COLORS["insight_bg"]
            apply_body(ws, r, c, fill_color=cell_bg)

        ws.cell(row=r, column=1, value=exp_id)
        ws.cell(row=r, column=1).font = Font(name="Arial", size=10, color="555555")
        ws.cell(row=r, column=section_ranges["META"][0] + 4, value="Planning")
        ws.cell(row=r, column=section_ranges["META"][0] + 4).font = Font(
            name="Arial", size=10, color="FF8F00", italic=True
        )

        # Δ vs Baseline 公式 (对比第一个 primary metric)
        baseline_cell = f"${get_column_letter(primary_start)}${baseline_row}"
        current_cell = f"{get_column_letter(primary_start)}{r}"
        formula = f'=IF(AND({current_cell}<>"",{baseline_cell}<>""),{current_cell}-{baseline_cell},"")'
        delta_cell = ws.cell(row=r, column=delta_col, value=formula)
        delta_cell.alignment = CENTER_ALIGN

    # ── 条件格式 ──
    last_row = baseline_row + num_rows
    surprise_range = f"{get_column_letter(surprise_col)}3:{get_column_letter(surprise_col)}{last_row}"
    ws.conditional_formatting.add(
        surprise_range,
        CellIsRule(operator="equal", formula=['"Yes"'],
                   fill=PatternFill("solid", fgColor="FFCC80"),
                   font=Font(bold=True, color=COLORS["surprise_yes"]))
    )
    ws.conditional_formatting.add(
        surprise_range,
        CellIsRule(operator="equal", formula=['"No"'],
                   fill=PatternFill("solid", fgColor="C8E6C9"),
                   font=Font(color=COLORS["surprise_no"]))
    )

    # Status 条件格式
    status_col = section_ranges["META"][0] + 4
    status_range = f"{get_column_letter(status_col)}3:{get_column_letter(status_col)}{last_row}"
    status_formats = {
        "Done": ("C8E6C9", "2E7D32"),
        "Running": ("BBDEFB", "1565C0"),
        "Failed": ("FFCDD2", "C62828"),
        "Abandoned": ("F5F5F5", "9E9E9E"),
    }
    for status, (fill_c, font_c) in status_formats.items():
        ws.conditional_formatting.add(
            status_range,
            CellIsRule(operator="equal", formula=[f'"{status}"'],
                       fill=PatternFill("solid", fgColor=fill_c),
                       font=Font(color=font_c, bold=True))
        )

    # Delta 条件格式
    delta_range = f"{get_column_letter(delta_col)}4:{get_column_letter(delta_col)}{last_row}"
    ws.conditional_formatting.add(
        delta_range,
        CellIsRule(operator="greaterThan", formula=["0"],
                   font=Font(color=COLORS["delta_positive"], bold=True))
    )
    ws.conditional_formatting.add(
        delta_range,
        CellIsRule(operator="lessThan", formula=["0"],
                   font=Font(color=COLORS["delta_negative"], bold=True))
    )

    # 冻结窗格
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(total_cols)}{last_row}"

    return ws


def create_exploration_sheet(wb, domain):
    """探索性实验 sheet"""
    ws = wb.create_sheet("Exploration")
    ws.sheet_properties.tabColor = "FF8F00"

    headers = [
        ("Exp ID", 10), ("Date", 12), ("Idea / Motivation", 35),
        ("What I Tried", 35), ("Result Summary", 25),
        ("Worth Pursuing?", 14), ("Promoted to Main?", 14),
        ("Notes", 35),
    ]
    for i, (name, width) in enumerate(headers):
        apply_header(ws, 1, i + 1, name, width)

    for r in range(2, 22):
        for c in range(1, len(headers) + 1):
            apply_body(ws, r, c)
        ws.cell(row=r, column=1, value=f"X{r-1:03d}")
        ws.cell(row=r, column=1).font = Font(name="Arial", size=10, color="888888")

    ws.freeze_panes = "A2"


def create_failed_sheet(wb):
    """失败/废弃实验 sheet"""
    ws = wb.create_sheet("Failed & Abandoned")
    ws.sheet_properties.tabColor = "C62828"

    headers = [
        ("Exp ID", 10), ("Date", 12), ("Original Description", 30),
        ("Failure Reason", 30), ("Error / Bug Detail", 30),
        ("Lesson Learned", 30), ("Avoid in Future", 25),
    ]
    for i, (name, width) in enumerate(headers):
        apply_header(ws, 1, i + 1, name, width)
        for r in range(2, 12):
            apply_body(ws, r, i + 1, fill_color=COLORS["failed_bg"])

    ws.freeze_panes = "A2"


def create_config_log(wb, exp_type):
    """配置日志 sheet"""
    ws = wb.create_sheet("Config Log")
    ws.sheet_properties.tabColor = "546E7A"

    headers = [
        ("Exp ID", 10), ("Config File Path", 30), ("Git Commit / Branch", 22),
        ("GPU Type", 14), ("GPU Count", 10), ("Random Seed", 10),
        ("Epochs", 10), ("Total Train Time", 14),
        ("Checkpoint Path", 30), ("Reproducibility Notes", 30),
    ]
    for i, (name, width) in enumerate(headers):
        apply_header(ws, 1, i + 1, name, width)
        for r in range(2, 22):
            apply_body(ws, r, i + 1, fill_color=COLORS["meta_bg"])

    ws.freeze_panes = "A2"


def create_notes_sheet(wb):
    """Notes & Insights sheet"""
    ws = wb.create_sheet("Notes & Insights")
    ws.sheet_properties.tabColor = "7B1FA2"

    headers = [
        ("Date", 12), ("Phase / Week", 14), ("Summary of Progress", 40),
        ("Key Insight", 40), ("Direction Change?", 14),
        ("Next Phase Plan", 35),
    ]
    for i, (name, width) in enumerate(headers):
        apply_header(ws, 1, i + 1, name, width)
        for r in range(2, 15):
            apply_body(ws, r, i + 1)

    ws.freeze_panes = "A2"


def generate_tracker(domain, exp_type, output_path, project_name, baseline_name, num_rows):
    wb = Workbook()

    create_dashboard(wb, project_name, domain, exp_type)
    create_main_experiments(wb, domain, exp_type, baseline_name, num_rows)
    create_exploration_sheet(wb, domain)
    create_failed_sheet(wb)
    create_config_log(wb, exp_type)
    create_notes_sheet(wb)

    wb.save(output_path)
    print(f"Experiment tracker saved to: {output_path}")
    return output_path


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Generate FAIR-style experiment tracker")
    parser.add_argument("--domain", default="general",
                        choices=["cv", "nlp", "rl", "multimodal", "general"])
    parser.add_argument("--experiment_type", default="ablation",
                        choices=["ablation", "hyperparam", "architecture", "method", "dataset"])
    parser.add_argument("--output", default="/mnt/user-data/outputs/experiment_tracker.xlsx")
    parser.add_argument("--project_name", default="My Research Project")
    parser.add_argument("--baseline_name", default="Baseline")
    parser.add_argument("--num_experiment_rows", type=int, default=20)

    args = parser.parse_args()
    generate_tracker(
        args.domain, args.experiment_type, args.output,
        args.project_name, args.baseline_name, args.num_experiment_rows
    )
